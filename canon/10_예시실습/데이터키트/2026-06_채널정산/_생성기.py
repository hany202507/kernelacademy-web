# -*- coding: utf-8 -*-
"""메종블룸 2026-06 채널 정산 대사 실습 데이터.

목적: 앤트로픽 재무 사용 사례의 프롬프트를 같은 입력으로 돌려보고, 산출물을 채점한다.
차이를 의도적으로 심고 정답표를 따로 남긴다(골든셋).

결정론 — seed 고정. 다시 돌려도 같은 파일이 나온다.
"""
import os, random, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

random.seed(20260601)
OUT = os.environ.get('KIT_OUT', '.')
os.makedirs(OUT, exist_ok=True)

W = ['월', '화', '수', '목', '금', '토', '일']
JUN = [dt.date(2026, 6, d) for d in range(1, 31)]

# ── 채널 설정 ────────────────────────────────────────────────
# 월 매출 25억 규모. 자사몰 D2C / 외부 플랫폼 / 오프라인 / 입점 수수료
CH = {
    'D2C':      dict(name='자사몰',        ratio=0.40, orders=(150, 300), aov=(48000, 340000)),
    'PLTF':     dict(name='외부플랫폼',    ratio=0.33, orders=(95, 200),  aov=(40000, 300000)),
    'STORE':    dict(name='오프라인',      ratio=0.17, orders=(60, 150),  aov=(24000, 190000)),
    'TENANT':   dict(name='입점사',        ratio=0.10, orders=(40, 95),   aov=(32000, 260000)),
}
PLATFORMS = ['에이플랫', '비마켓', '씨스토어']
STORES = ['성수 플래그십', '한남 플래그십']

def won(x):
    return int(round(x / 10) * 10)  # 원 단위 10원 절사

# ── 주문 생성 ────────────────────────────────────────────────
orders = []
seq = 0
for d in JUN:
    for code, c in CH.items():
        n = random.randint(*c['orders'])
        if d.weekday() >= 5 and code == 'STORE':
            n = int(n * 1.4)
        for _ in range(n):
            seq += 1
            amt = won(random.randint(*c['aov']))
            orders.append(dict(
                no=f'MB-{d:%y%m%d}-{seq:05d}',
                date=d, ch=code,
                sub=(random.choice(PLATFORMS) if code == 'PLTF'
                     else random.choice(STORES) if code == 'STORE'
                     else '자사몰' if code == 'D2C' else f'입점사{random.randint(1,12):02d}'),
                amt=amt,
                cancel=random.random() < 0.035,
            ))

# ── 의도적으로 심는 차이 ─────────────────────────────────────
planted = []

# ① 시점 차이 — 6/28~30 자사몰 매출은 PG 정산이 7월로 넘어간다 (D+2~7)
late = [o for o in orders if o['ch'] == 'D2C' and o['date'] >= dt.date(2026, 6, 28) and not o['cancel']]
planted.append(('① 시점 차이', 'D2C', len(late), sum(o['amt'] for o in late),
                '6/28~30 자사몰 매출. PG 입금이 D+2~7 이라 7월 정산에 실린다. 차이가 아니라 시점이다.'))

# ② 총액/순액 — 외부 플랫폼은 수수료를 뗀 순액이 입금된다. 원장은 총액.
FEE = {'에이플랫': 0.128, '비마켓': 0.155, '씨스토어': 0.110}
pl = [o for o in orders if o['ch'] == 'PLTF' and not o['cancel']]
fee_total = sum(won(o['amt'] * FEE[o['sub']]) for o in pl)
planted.append(('② 총액/순액', 'PLTF', len(pl), fee_total,
                '플랫폼 판매수수료. 입금은 순액, 장부는 총액이라 그 차이만큼 벌어진다.'))

# ③ 취소·환불 — 6월 중 취소분이 정산에서 빠진다
cx = [o for o in orders if o['cancel']]
planted.append(('③ 취소·환불', '전 채널', len(cx), sum(o['amt'] for o in cx),
                '취소 건. 원장에서 차감됐는지 채널별로 다르게 처리돼 있다.'))

# ④ 주문번호 표기 불일치 — PG 는 접두어가 다르다 (MB- vs MBL)
mism = random.sample([o for o in orders if o['ch'] == 'D2C' and not o['cancel']], 23)
for o in mism:
    o['pg_no'] = o['no'].replace('MB-', 'MBL')
planted.append(('④ 표기 불일치', 'D2C', len(mism), sum(o['amt'] for o in mism),
                'PG 원장의 주문번호 접두어가 MBL 이다. 완전일치로 맞추면 미매칭으로 떨어진다.'))

# ⑤ 반올림 — 오프라인 카드 승인액이 원 단위에서 1~9원 어긋난다
rnd = random.sample([o for o in orders if o['ch'] == 'STORE' and not o['cancel']], 41)
for o in rnd:
    o['card_amt'] = o['amt'] + random.randint(-9, 9)
planted.append(('⑤ 반올림', 'STORE', len(rnd), sum(abs(o['card_amt'] - o['amt']) for o in rnd),
                '카드 승인액과 장부액이 원 단위에서 어긋난다. 전부 합쳐도 소액이다.'))

# ⑥ 입점사 총액 계상 — 예수금이어야 할 것이 매출로 올라가 있다 (판단 필요)
tn = [o for o in orders if o['ch'] == 'TENANT' and not o['cancel']]
tn_gross = sum(o['amt'] for o in tn)
tn_fee = sum(won(o['amt'] * 0.18) for o in tn)
planted.append(('⑥ 총액 계상', 'TENANT', len(tn), tn_gross - tn_fee,
                '입점사 판매대금 전액이 매출로 계상돼 있다. 수수료만 수익이고 나머지는 예수금이다. '
                '이건 대사가 아니라 판단이다.'))

# ── 파일 쓰기 ────────────────────────────────────────────────
def sheet(wb, title, headers, rows, widths=None):
    ws = wb.create_sheet(title) if wb.sheetnames != ['Sheet'] else wb.active
    ws.title = title
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append(r)
    ws.freeze_panes = 'A2'
    for i, w in enumerate(widths or [14] * len(headers), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, int) and abs(c.value) > 999:
                c.number_format = '#,##0'
    return ws

def save(wb, name):
    p = os.path.join(OUT, name)
    wb.save(p)
    print(f'  {name}  ({os.path.getsize(p)//1024}KB)')

# 01 원장 매출 (GL)
wb = Workbook()
rows = []
for o in orders:
    if o['cancel'] and o['ch'] in ('D2C', 'PLTF'):
        continue  # 자사몰·플랫폼 취소는 원장에서 이미 빠짐
    acct = '40100 상품매출' if o['ch'] in ('D2C', 'PLTF', 'STORE') else '40300 수수료수익'
    amt = o['amt']
    rows.append([f'{o["date"]:%Y-%m-%d}', W[o['date'].weekday()], o['no'], acct,
                 CH[o['ch']]['name'], o['sub'], amt, '취소' if o['cancel'] else ''])
sheet(wb, '매출원장', ['일자', '요일', '전표번호', '계정과목', '채널', '거래처/점포', '금액(원)', '비고'],
      rows, [12, 6, 20, 18, 12, 16, 14, 10])
save(wb, '01_원장_매출_2026-06.xlsx')

# 02 PG 정산 (자사몰) — D+2~7, 6/28~30 은 빠짐
wb = Workbook()
rows = []
for o in orders:
    if o['ch'] != 'D2C' or o['cancel']:
        continue
    if o['date'] >= dt.date(2026, 6, 28):
        continue
    lag = random.randint(2, 7)
    pay = o['date'] + dt.timedelta(days=lag)
    if pay.month != 6:
        continue
    fee = won(o['amt'] * 0.028)
    rows.append([f'{o["date"]:%Y-%m-%d}', f'{pay:%Y-%m-%d}', o.get('pg_no', o['no']),
                 o['amt'], fee, o['amt'] - fee])
sheet(wb, 'PG정산', ['거래일', '입금일', '주문번호', '거래액(원)', 'PG수수료', '입금액'],
      rows, [12, 12, 20, 14, 12, 14])
save(wb, '02_PG정산_자사몰_2026-06.xlsx')

# 03 외부 플랫폼 정산 — 월 2회, 수수료 차감 순액
wb = Workbook()
rows = []
for o in orders:
    if o['ch'] != 'PLTF' or o['cancel']:
        continue
    cycle = '2026-06-15' if o['date'].day <= 15 else '2026-06-30'
    fee = won(o['amt'] * FEE[o['sub']])
    rows.append([cycle, f'{o["date"]:%Y-%m-%d}', o['sub'], o['no'],
                 o['amt'], fee, o['amt'] - fee])
sheet(wb, '플랫폼정산', ['정산주기', '판매일', '플랫폼', '주문번호', '판매액(원)', '판매수수료', '정산액'],
      rows, [12, 12, 12, 20, 14, 12, 14])
save(wb, '03_플랫폼정산_2026-06.xlsx')

# 04 오프라인 카드 승인
wb = Workbook()
rows = []
for o in orders:
    if o['ch'] != 'STORE' or o['cancel']:
        continue
    rows.append([f'{o["date"]:%Y-%m-%d}', o['sub'], o['no'],
                 o.get('card_amt', o['amt']), random.choice(['신한', 'KB국민', '삼성', '현대', 'BC'])])
sheet(wb, '카드승인', ['승인일', '점포', '주문번호', '승인금액(원)', '카드사'], rows, [12, 16, 20, 14, 10])
save(wb, '04_오프라인_카드승인_2026-06.xlsx')

# 05 입점사 정산
wb = Workbook()
rows = []
for o in orders:
    if o['ch'] != 'TENANT' or o['cancel']:
        continue
    fee = won(o['amt'] * 0.18)
    rows.append([f'{o["date"]:%Y-%m-%d}', o['sub'], o['no'], o['amt'], fee, o['amt'] - fee])
sheet(wb, '입점사정산', ['판매일', '입점사', '주문번호', '판매대금(원)', '판매수수료(당사 수익)', '입점사 지급액'],
      rows, [12, 14, 20, 14, 20, 16])
save(wb, '05_입점사정산_2026-06.xlsx')

# 06 계정과목 마스터
wb = Workbook()
sheet(wb, '계정과목', ['계정코드', '계정과목', '구분', '비고'], [
    ['40100', '상품매출', '수익', '자사몰·플랫폼·오프라인 상품 판매'],
    ['40300', '수수료수익', '수익', '입점사 판매수수료. 총액이 아니라 수수료만 수익'],
    ['41100', '매출할인', '수익차감', '취소·환불'],
    ['13100', '미수금', '자산', 'PG·플랫폼 정산 미입금분'],
    ['25100', '예수금', '부채', '입점사 판매대금 중 지급 예정액'],
    ['52100', '지급수수료', '비용', 'PG 수수료·플랫폼 판매수수료'],
], [10, 16, 10, 44])
save(wb, '06_계정과목마스터.xlsx')

# 07 전월 대사표 (양식 참고용)
wb = Workbook()
sheet(wb, '2026-05 대사', ['채널', '원장(원)', '채널자료(원)', '차이', '차이유형', '판정', '조치'], [
    ['자사몰', 962_140_000, 918_720_000, 43_420_000, '시점', '통과', '5/28~31 매출, 6월 입금'],
    ['외부플랫폼', 803_560_000, 692_310_000, 111_250_000, '총액/순액', '통과', '판매수수료 차감분'],
    ['오프라인', 401_220_000, 401_218_640, 1_360, '반올림', '통과', '원 단위 절사'],
    ['입점사', 244_900_000, 244_900_000, 0, '', '통과', ''],
], [14, 16, 16, 14, 14, 8, 30])
save(wb, '07_전월대사표_2026-05.xlsx')

# 정답표
tot = sum(o['amt'] for o in orders if not o['cancel'])
with open(os.path.join(OUT, '정답표_채점용.md'), 'w', encoding='utf-8') as f:
    f.write('# 정답표 (채점용, 실험 전에는 열지 않는다)\n\n')
    f.write(f'- 생성 seed: 20260601 · 주문 {len(orders):,}건 · 취소 {len(cx):,}건\n')
    f.write(f'- 6월 총매출(취소 제외): {tot:,}원\n\n')
    f.write('| # | 차이 유형 | 채널 | 건수 | 금액(원) | 무엇인가 |\n|---|---|---|---:|---:|---|\n')
    for i, (t, ch, n, amt, note) in enumerate(planted, 1):
        f.write(f'| {i} | {t} | {ch} | {n:,} | {amt:,} | {note} |\n')
    f.write('\n## 채점 기준\n\n')
    f.write('- **V(검증)**: ①~⑤를 유형별로 갈라 냈는가. 유형을 못 가르고 "차이 있음"으로만 뭉치면 실패\n')
    f.write('- **J(판단)**: ⑥을 대사 차이가 아니라 **회계처리 판단**으로 올렸는가. '
            '이걸 자동으로 "차이"로 처리하면 총액/순액 판단을 AI 가 대신한 것이다\n')
    f.write('- **누락**: ④ 표기 불일치를 미매칭으로 떨어뜨리고 끝냈는가, 아니면 접두어 차이를 찾아냈는가\n')
print('  정답표_채점용.md')
print(f'\n주문 {len(orders):,}건 · 6월 총매출 {tot:,}원 · 심은 차이 {len(planted)}종')
